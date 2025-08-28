from datetime import datetime, timedelta
from io import BytesIO
from typing import Dict, List
import calendar

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------
# Legends (codes & meanings)
# --------------------------
REGULAR_LEGEND: Dict[str, str] = {
    "P": "Present",
    "A": "Absent",
    "R": "Rest",
    "CL": "Casual Leave",
    "LAP": "Leave On Average Pay",
    "COCL": "Compensatory Casual Leave",
    "S": "Sick",
    "ART": "Accident Relief Train",
    "Trg": "Training",
    "SCL": "Special Casual Leave",
    "H": "Holiday",
    "D": "Duty",
    "Ex": "Exam",
    "Sp": "Spare",
    "Trans": "Transfer",
    "Retd": "Retired",
    "Rel": "Released",
}

APPRENTICE_LEGEND: Dict[str, str] = {
    "P": "Present",
    "A": "Absent",
    "R": "Rest",
    "S": "Sick",
    "CL": "Casual Leave",
    "REL": "Released",
}

# --------------------------
# Styles & fills
# --------------------------
TITLE_FONT = Font(bold=True, size=16)
HEADER_FONT = Font(bold=True, size=14)
BOLD = Font(bold=True, size=12)
NORMAL = Font(size=12)
SMALL_ITALIC = Font(italic=True, size=10, bold=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
JUSTIFY = Alignment(horizontal="justify", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# Header shading
HEADER_SHADE = PatternFill("solid", fgColor="EEEEEE")

# Calendar backgrounds
SUNDAY_FILL = PatternFill("solid", fgColor="E6E6E6")   # light gray
HOLIDAY_FILL = PatternFill("solid", fgColor="F4CCCC")  # light pink

# Optional fills per attendance code
CODE_FILLS: Dict[str, PatternFill] = {
    "A": PatternFill("solid", fgColor="FFF2CC"),  # pale yellow
    "R": PatternFill("solid", fgColor="D9EAD3"),  # pale green
    "S": PatternFill("solid", fgColor="D0E0E3"),  # pale blue
    "CL": PatternFill("solid", fgColor="FCE5CD"), # peach
    "LAP": PatternFill("solid", fgColor="FCE5CD"),
    "COCL": PatternFill("solid", fgColor="FCE5CD"),
    "H": HOLIDAY_FILL,
}

WEEKDAYS_ABBR = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def _date_range(employee_type: str, month: str) -> List[datetime]:
    """Return list of datetime days in range based on employee type."""
    year, mon = map(int, month.split("-"))
    if employee_type.lower() == "regular":
        start_date = datetime(year, mon, 11)
        if mon == 12:
            end_date = datetime(year + 1, 1, 10)
        else:
            end_date = datetime(year, mon + 1, 10)
    else:
        start_date = datetime(year, mon, 1)
        end_date = datetime(year, mon, calendar.monthrange(year, mon)[1])
    total_days = (end_date - start_date).days + 1
    return [start_date + timedelta(days=i) for i in range(total_days)]


async def create_attendance_excel(db, employee_type: str, month: str) -> BytesIO:
    """Build a calendar-styled attendance workbook."""
    # --------------------------
    # Fetch calendar & data
    # --------------------------
    dates = _date_range(employee_type, month)
    legend = REGULAR_LEGEND if employee_type.lower() == "regular" else APPRENTICE_LEGEND

    emp_cursor = db["employees"].find({"type": employee_type})
    employees = [emp async for emp in emp_cursor]

    start_s, end_s = dates[0].strftime("%Y-%m-%d"), dates[-1].strftime("%Y-%m-%d")
    hol_cursor = db["holidays"].find({"date": {"$gte": start_s, "$lte": end_s}})
    holidays = {doc["date"] async for doc in hol_cursor}

    att_cursor = db["attendance"].find({"type": employee_type, "month": month})
    attendance = {doc["emp_no"]: doc["records"] async for doc in att_cursor}

    # --------------------------
    # Workbook & header section
    # --------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    fixed_cols = 4
    total_cols = fixed_cols + len(dates)

    # Title rows
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"] = "SOUTH EASTERN RAILWAY"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws["A2"] = "ELECTRICAL DEPARTMENT"
    ws["A2"].font = HEADER_FONT
    ws["A2"].alignment = CENTER

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
    ws["A3"] = "OFFICE OF THE SENIOR SECTION ENGINEER (ELECT.)/SW/KGP"
    ws["A3"].font = HEADER_FONT
    ws["A3"].alignment = CENTER

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_cols)
    ws["A4"] = f"ATTENDANCE SHEET / MUSTER ROLL ({month})"
    ws["A4"].font = TITLE_FONT
    ws["A4"].alignment = CENTER

    ws.append([])

    # --------------------------
    # Table header section
    # --------------------------
    header_row = ws.max_row + 1
    titles = ["S.No", "NAME", "DESIGNATION", "EMPLOYEE NO."] + [d.strftime("%d/%m") for d in dates]
    ws.append(titles)
    ws.append(["", "", "", ""] + [WEEKDAYS_ABBR[d.weekday()] for d in dates])

    for col, label in enumerate(["S.No", "NAME", "DESIGNATION", "EMPLOYEE NO."], start=1):
        ws.merge_cells(start_row=header_row, start_column=col,
                       end_row=header_row + 1, end_column=col)
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = BOLD
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.fill = HEADER_SHADE

    for offset, d in enumerate(dates):
        col = fixed_cols + 1 + offset
        date_cell = ws.cell(row=header_row, column=col)
        day_cell = ws.cell(row=header_row + 1, column=col)
        date_cell.font = BOLD
        day_cell.font = Font(bold=True, size=10)
        date_cell.alignment, day_cell.alignment = CENTER, CENTER
        date_cell.border, day_cell.border = THIN_BORDER, THIN_BORDER
        if d.weekday() == 6:
            date_cell.fill = SUNDAY_FILL
            day_cell.fill = SUNDAY_FILL
        if d.strftime("%Y-%m-%d") in holidays:
            date_cell.fill = HOLIDAY_FILL
            day_cell.fill = HOLIDAY_FILL

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    for col in range(fixed_cols + 1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    data_start_row = header_row + 2
    ws.freeze_panes = f"E{data_start_row}"

    # --------------------------
    # Employee rows
    # --------------------------
    row = data_start_row
    for idx, emp in enumerate(employees, start=1):
        ws.row_dimensions[row].height = 20  # <<< row height
        ws.cell(row=row, column=1, value=idx).alignment = CENTER
        ws.cell(row=row, column=2, value=emp.get("name", "")).font = NORMAL
        ws.cell(row=row, column=3, value=emp.get("designation", "")).font = NORMAL
        ws.cell(row=row, column=4, value=emp.get("emp_no", "")).font = NORMAL

        emp_att = attendance.get(emp.get("emp_no", ""), {})
        for off, d in enumerate(dates):
            col = fixed_cols + 1 + off
            date_str = d.strftime("%Y-%m-%d")
            code_val = emp_att.get(date_str, "")
            code_key = str(code_val).split("/")[0] if code_val else ""
            cell = ws.cell(row=row, column=col, value=code_val)
            cell.alignment, cell.font, cell.border = CENTER, NORMAL, THIN_BORDER
            if d.weekday() == 6:
                cell.fill = SUNDAY_FILL
            if date_str in holidays:
                cell.fill = HOLIDAY_FILL
            if code_key in CODE_FILLS:
                cell.fill = CODE_FILLS[code_key]
        for c in range(1, fixed_cols + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    last_data_row = row - 1

    # --------------------------
    # Legends + Note + Signatures
    # --------------------------
    row = last_data_row + 2
    legend_box_start_col, legend_box_end_col = 2, 12

    # LEGENDS title
    ws.merge_cells(start_row=row, start_column=legend_box_start_col,
                   end_row=row, end_column=legend_box_end_col)
    tcell = ws.cell(row=row, column=legend_box_start_col, value="LEGENDS")
    tcell.font = Font(bold=True, size=12)
    tcell.alignment = CENTER
    tcell.fill = HEADER_SHADE

    # Create single paragraph with all legend codes
    row += 1
    legend_codes_text = ", ".join([f"{k} = {v}" for k, v in legend.items()])
    
    ws.row_dimensions[row].height = 40  # Adequate height for wrapped text
    ws.merge_cells(start_row=row, start_column=legend_box_start_col,
                   end_row=row, end_column=legend_box_end_col)
    codes_cell = ws.cell(row=row, column=legend_box_start_col, value=legend_codes_text)
    codes_cell.font = Font(size=11)
    codes_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Note section
    row += 1
    ws.row_dimensions[row].height = 35
    ws.merge_cells(start_row=row, start_column=legend_box_start_col,
                   end_row=row, end_column=legend_box_end_col)
    ncell = ws.cell(row=row, column=legend_box_start_col, value=(
        "Note: The above abstract attendance particulars are taken from the "
        "attendance register for staff of O/O SSEE/SW/KGP. Due to unavoidable "
        "circumstances, manual entries may have been made by the signatory."
    ))
    
    ncell.font = SMALL_ITALIC
    ncell.alignment = JUSTIFY

    # Apply borders to the entire legends box
    legend_title_row = last_data_row + 3
    block_end_row = row
    for r in range(legend_title_row, block_end_row + 1):
        for c in range(legend_box_start_col, legend_box_end_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

    # Signature section
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=legend_box_end_col + 8)
    ws.cell(row=row, column=1, value=(
        "JE: ______________      SSEE: ______________      SSE/INCHARGE: ______________"
    )).alignment = CENTER

    # --------------------------
    # Page setup
    # --------------------------
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = f"{header_row}:{header_row+1}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output